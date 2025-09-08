# main.py (CAIO Backend) — drop-in replacement
import os, logging, traceback, json, re
from typing import Optional, List, Tuple
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Request, Body, Query
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import func

from db import get_db, User, init_db, UsageLog, ChatSession, ChatMessage
from auth import create_access_token, verify_password, get_password_hash, get_current_user

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("caio")
DEBUG = os.getenv("DEBUG", "0") in ("1","true","TRUE","yes","YES")

app = FastAPI(title="CAIO Backend", version="0.3.1")

@app.get("/")
def index():
    return {"ok": True, "service": "caio-backend", "time": datetime.utcnow().isoformat() + "Z"}

@app.on_event("startup")
def _startup_init_db():
    try:
        from time import sleep
        attempts = 0
        while attempts < 5:
            try:
                init_db()
                break
            except Exception as e:
                attempts += 1
                logging.warning(f"init_db() failed (attempt {attempts}/5): {e}")
                sleep(2 * attempts)
        else:
            logging.error("init_db() failed after retries; continuing to serve /api/health")
    except Exception as e:
        logging.error(f"startup hook error: {e}")

# ---------- CORS ----------
ALLOWED_ORIGINS = [
    "https://caio-frontend.vercel.app",
    "https://caioai.netlify.app",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
]
extra = os.getenv("ALLOWED_ORIGINS", "")
if extra:
    ALLOWED_ORIGINS += [o.strip() for o in extra.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
    max_age=86400,
)

@app.options("/{path:path}")
def cors_preflight(path: str):
    return JSONResponse({"ok": True})

# ---------- Health ----------
@app.get("/api/health")
def health():
    return {"status": "ok", "version": "0.3.1"}

@app.get("/api/ready")
def ready():
    return {"ready": True, "time": datetime.utcnow().isoformat() + "Z"}

# ---------- Auth / Tiers ----------
ADMIN_EMAILS = {e.strip().lower() for e in os.getenv("ADMIN_EMAILS", "vineetpjoshi.71@gmail.com").split(",") if e.strip()}
PREMIUM_EMAILS = {e.strip().lower() for e in os.getenv("PREMIUM_EMAILS", "").split(",") if e.strip()}

# NEW: honor both legacy and new env names seamlessly
DEMO_DAILY_LIMIT = int(os.getenv("DEMO_DAILY_LIMIT", os.getenv("FREE_QUERIES_PER_DAY", "5")))
PRO_DAILY_LIMIT  = int(os.getenv("PRO_DAILY_LIMIT",  os.getenv("PRO_QUERIES_PER_DAY", "50")))

def _login_core(email: str, password: str, db: Session):
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password required")
    user = db.query(User).filter(User.email == email).first()
    if user:
        if not hasattr(user, "hashed_password"):
            raise RuntimeError("User model missing 'hashed_password'")
        if not verify_password(password, user.hashed_password):
            raise HTTPException(status_code=400, detail="Incorrect email or password")
        # auto-sync flags from env on every login
        email_l = email.lower()
        is_admin_now = email_l in ADMIN_EMAILS
        is_premium_now = email_l in PREMIUM_EMAILS
        changed = False
        if hasattr(user, "is_admin") and user.is_admin != is_admin_now:
            user.is_admin = is_admin_now
            changed = True
        if hasattr(user, "is_paid") and (is_admin_now or is_premium_now) and not user.is_paid:
            user.is_paid = True
            changed = True
        if changed:
            db.add(user); db.commit(); db.refresh(user)
    else:
        hpw = get_password_hash(password)
        user = User(
            email=email,
            hashed_password=hpw if hasattr(User, "hashed_password") else None,
            is_admin=(email in ADMIN_EMAILS) if hasattr(User, "is_admin") else False,
            is_paid=(email in ADMIN_EMAILS or email in PREMIUM_EMAILS) if hasattr(User, "is_paid") else False,
            created_at=datetime.utcnow() if hasattr(User, "created_at") else None,
        )
        db.add(user); db.commit(); db.refresh(user)
    token = create_access_token(sub=user.email)
    return {
        "access_token": token,
        "token_type": "bearer",
        "email": user.email,
        "is_admin": bool(getattr(user, "is_admin", False)),
        "is_paid": bool(getattr(user, "is_paid", False)),
    }

@app.post("/api/login")
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    try:
        return _login_core((form.username or "").strip().lower(), form.password or "", db)
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Login failed: %s\n%s", e, traceback.format_exc())
        if DEBUG:
            return JSONResponse({"detail": f"login-500: {e}"}, status_code=500)
        raise HTTPException(status_code=500, detail="Login failed")

@app.get("/api/profile")
def profile(current_user: User = Depends(get_current_user)):
    email = current_user.email.lower()
    tier = "admin" if email in ADMIN_EMAILS else (
        "premium" if email in PREMIUM_EMAILS else (
            "pro" if getattr(current_user, "is_paid", False) else "demo"
        )
    )
    return {
        "email": current_user.email,
        "is_admin": email in ADMIN_EMAILS,
        "is_paid": bool(getattr(current_user, "is_paid", False)),
        "created_at": getattr(current_user, "created_at", None),
        "tier": tier,
    }

# ---------- Limits ----------
def _user_tier(user: User) -> str:
    email = (user.email or "").lower()
    if email in ADMIN_EMAILS: return "admin"
    if email in PREMIUM_EMAILS: return "premium"
    return "pro" if bool(getattr(user, "is_paid", False)) else "demo"

def _limit_for_tier(tier: str) -> Optional[int]:
    if tier in ("admin", "premium"): return None
    return PRO_DAILY_LIMIT if tier == "pro" else DEMO_DAILY_LIMIT

def _today_range_utc() -> Tuple[datetime, datetime]:
    now = datetime.utcnow()
    return datetime(now.year, now.month, now.day), datetime(now.year, now.month, now.day, 23, 59, 59, 999999)

def _check_and_increment_usage(db: Session, user: User, endpoint: str = "analyze"):
    tier = _user_tier(user)
    limit = _limit_for_tier(tier)
    start, end = _today_range_utc()
    if limit is None:
        db.add(UsageLog(user_id=getattr(user, "id", 0) or 0, timestamp=datetime.utcnow(), endpoint=endpoint, status="ok"))
        db.commit()
        return True, 0, None, end, tier
    used = (
        db.query(UsageLog)
        .filter(UsageLog.user_id == getattr(user, "id", 0))
        .filter(UsageLog.endpoint == endpoint)
        .filter(UsageLog.timestamp >= start)
        .filter(UsageLog.timestamp <= end)
        .count()
    )
    if used >= limit:
        return False, used, limit, end, tier
    db.add(UsageLog(user_id=getattr(user, "id", 0) or 0, timestamp=datetime.utcnow(), endpoint=endpoint, status="ok"))
    db.commit()
    return True, used + 1, limit, end, tier

def _limit_response(used: int, limit: int, reset_at: datetime, tier: str):
    plan = "Pro" if tier == "pro" else "Demo"
    return JSONResponse(
        status_code=429,
        content={
            "status": "error",
            "title": "Daily limit reached",
            "message": f"You've used {used}/{limit} {plan} requests today. Resets at {reset_at.strftime('%H:%M')} UTC.",
            "plan": tier,
            "used": used,
            "limit": limit,
            "remaining": max(0, limit - used),
            "reset_at": reset_at.isoformat() + "Z",
        },
    )

# ---------- File parsing ----------
def _read_txt(body: bytes) -> str:
    try:
        return body.decode("utf-8", errors="ignore")
    except Exception:
        return ""

def _read_pdf(body: bytes) -> str:
    try:
        from PyPDF2 import PdfReader
        return "\n".join((p.extract_text() or "") for p in PdfReader(BytesIO(body)).pages)
    except Exception:
        return ""

def _read_docx(body: bytes) -> str:
    try:
        import docx
        return "\n".join(p.text for p in docx.Document(BytesIO(body)).paragraphs)
    except Exception:
        return ""

def _read_csv(body: bytes) -> str:
    try:
        return "\n".join(body.decode("utf-8", errors="ignore").splitlines()[:200])
    except Exception:
        return ""

def _read_xlsx(body: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(body), data_only=True)
        out = []
        for ws in wb.worksheets[:2]:
            out.append(f"# Sheet: {ws.title}")
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i > 200: break
                cells = ["" if c is None else str(c) for c in row]
                out.append(",".join(cells))
        return "\n".join(out)
    except Exception:
        return ""

async def _extract_text(file: Optional[UploadFile]) -> str:
    if not file:
        return ""
    body = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".pdf"):  return _read_pdf(body)
    if name.endswith(".docx"): return _read_docx(body)
    if name.endswith(".xlsx"): return _read_xlsx(body)
    if name.endswith(".csv"):  return _read_csv(body)
    return _read_txt(body)

# ---------- Analyze ----------
DEFAULT_BRAINS_ORDER = ["CFO","COO","CHRO","CMO","CPO"]

def _choose_brains(requested: Optional[str], is_paid: bool) -> List[str]:
    if requested:
        items = [b.strip().upper() for b in requested.split(",") if b.strip()]
        chosen = [b for b in items if b in DEFAULT_BRAINS_ORDER] or DEFAULT_BRAINS_ORDER[:]
    else:
        chosen = DEFAULT_BRAINS_ORDER[:]
    return chosen if is_paid else chosen[:2]

def _brain_prompt(brief: str, extracted: str, brain: str) -> str:
    role_map = {
        "CFO":"Chief Financial Officer — unit economics; revenue mix, margins, CCC, runway.",
        "COO":"Chief Operating Officer — cost-to-serve & reliability; capacity, throughput, SLA.",
        "CHRO":"Chief Human Resources Officer — org effectiveness; attrition, engagement.",
        "CMO":"Chief Marketing Officer — efficient growth; CAC/LTV, funnel, retention.",
        "CPO":"Chief People Officer — talent acquisition; pipeline, time-to-hire, QoH.",
    }
    role = role_map.get(brain, "Executive Advisor")
    return f"""
You are {role}.
Return STRICT MARKDOWN ONLY for **{brain}** with this structure:

### Insights
1. <insight>
2. <insight>
3. <insight>

### Recommendations
1. **<headline>**: <ONE sentence action.>
2. **<headline>**: <ONE sentence action.>
3. **<headline>**: <ONE sentence action.>

[BRIEF]
{brief or "(none)"}

[DOCUMENT TEXT]
{(extracted or "")[:120000]}
""".strip()

def _pick_provider(is_paid: bool) -> List[str]:
    raw = os.getenv("LLM_PROVIDER_PRO" if is_paid else "LLM_PROVIDER_DEMO", "openai,openrouter" if is_paid else "openrouter")
    providers = [p.strip().lower() for p in raw.split(",") if p.strip()]
    available = []
    if "openai" in providers and os.getenv("OPENAI_API_KEY"): available.append("openai")
    if "openrouter" in providers and os.getenv("OPENROUTER_API_KEY"): available.append("openrouter")
    return available or []

def _call_openai_prompt(prompt: str) -> str:
    key = os.getenv("OPENAI_API_KEY"); import requests
    headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    if os.getenv("OPENAI_ORG"): headers["OpenAI-Organization"] = os.getenv("OPENAI_ORG")
    if os.getenv("OPENAI_PROJECT"): headers["OpenAI-Project"] = os.getenv("OPENAI_PROJECT")
    data = {"model":"gpt-4o-mini","messages":[{"role":"user","content":prompt}],"temperature":0.2}
    r = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, data=json.dumps(data), timeout=60)
    if r.status_code >= 400: raise RuntimeError(f"OpenAI {r.status_code}: {r.text}")
    return r.json()["choices"][0]["message"]["content"]

def _call_openrouter_prompt(prompt: str) -> str:
    key = os.getenv("OPENROUTER_API_KEY"); import requests
    headers = {"Authorization": f"Bearer {key}","HTTP-Referer":"https://caio-frontend.vercel.app","X-Title":"CAIO","Content-Type":"application/json"}
    data = {"model":"openai/gpt-4o-mini","messages":[{"role":"user","content":prompt}],"temperature":0.2}
    r = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, data=json.dumps(data), timeout=60)
    if r.status_code >= 400: raise RuntimeError(f"OpenRouter {r.status_code}: {r.text}")
    return r.json()["choices"][0]["message"]["content"]

@app.post("/api/analyze")
async def analyze(
    request: Request,
    file: Optional[UploadFile] = File(None),
    text: Optional[str] = Form(None),
    brains: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    allowed, used, limit, reset_at, tier = _check_and_increment_usage(db, current_user, endpoint="analyze")
    if not allowed:
        return _limit_response(used, int(limit or 0), reset_at, tier)

    if tier == "demo":
        chosen = _choose_brains(brains, is_paid=False)
        return JSONResponse({
            "status":"demo",
            "title": f"Demo Mode · {', '.join(chosen)}",
            "summary":"This is a demo preview. Upgrade to Pro to run all brains with real analysis.\n"
                      f"Brains used: {', '.join(chosen)}",
            "tip":"Upload a business document or provide a brief to see the flow.",
        })

    try:
        extracted = await _extract_text(file)
    except Exception as e:
        logger.error("Extraction failed: %s", e)
        extracted = ""
    brief = (text or "").strip()

    # If you have a smarter ordering function elsewhere, use it; else fall back:
    try:
        chosen = order_brains_for_content(brief, extracted, is_paid=True) if not brains else _choose_brains(brains, True)  # type: ignore
    except Exception:
        chosen = _choose_brains(brains, True)

    if not extracted and not brief:
        raise HTTPException(status_code=400, detail="Please upload a file or provide text")

    prompts = {b: _brain_prompt(brief, extracted, b) for b in chosen}
    provider_chain = _pick_provider(is_paid=True)
    summaries, used_provider, errs = {}, "stub", []
    if provider_chain:
        for p in provider_chain:
            try:
                if p == "openai":
                    summaries = {b: _call_openai_prompt(pr) for b, pr in prompts.items()}
                elif p == "openrouter":
                    summaries = {b: _call_openrouter_prompt(pr) for b, pr in prompts.items()}
                used_provider = p
                break
            except Exception as e:
                logger.error("Provider %s failed: %s", p, e)
                errs.append(f"{p}:{e.__class__.__name__}")
                summaries = {}
    if not summaries:
        summaries = {b: f"[Stub after error] {b} analysis. Errors: {', '.join(errs) if errs else 'no-provider'}."
                     for b in prompts.keys()}
    combined = [f"### {b}\n{summaries[b]}" for b in chosen]
    return {
        "status":"ok",
        "title": f"Analysis Complete · {', '.join(chosen)}",
        "summary": "\n\n".join(combined),
        "meta": {"provider": used_provider, "brains": chosen, "chars": len(extracted)},
    }

# ---------- Premium Chat ----------
def _is_premium_or_admin(user: User) -> bool:
    email = (user.email or "").lower()
    return email in ADMIN_EMAILS or email in PREMIUM_EMAILS

# ====== PRIMARY CXO ROUTING ======
PRIMARY_KEYWORDS = {
    "CFO": [
        "revenue","sales","turnover","profit","margin","ebitda","p&l","pnl","balance sheet","cash flow",
        "runway","burn","budget","forecast","pricing","unit economics","cost of goods","cogs","ar","ap",
        "working capital","inventory","gross margin","net margin"
    ],
    "COO": [
        "operations","operational","throughput","capacity","utilization","sla","uptime","downtime","logistics",
        "supply chain","fulfilment","fulfillment","warehouse","production","quality","defect","lead time",
        "process","workflow","on-time","takt","bottleneck"
    ],
    "CHRO": [
        "attrition","churn of employees","engagement","morale","headcount","hiring","performance review",
        "compensation","pay","benefits","hr","policy","culture","training","learning","succession","l&d"
    ],
    "CMO": [
        "marketing","campaign","brand","awareness","lead","mql","sql","conversion","funnel","ad","seo","sem",
        "content","social","newsletter","cac","ltv","retention","churn","activation"
    ],
    "CPO": [
        "recruiting","talent","candidate","pipeline","sourcing","screening","time to hire","offer acceptance",
        "quality of hire","employer brand","interview"
    ],
}

MODEL_HINT_MAP = {
    "CFO": "Start with CFO Recommendations first (numbered). Reference metrics like revenue growth, margins, CAC/LTV, cash runway.",
    "COO": "Start with COO Recommendations first (numbered). Focus on throughput, reliability, process constraints, and cost-to-serve.",
    "CHRO": "Start with CHRO Recommendations first (numbered). Address attrition, engagement, hiring pipelines, and org effectiveness.",
    "CMO": "Start with CMO Recommendations first (numbered). Cover growth levers, funnel conversion, retention and ROI.",
    "CPO": "Start with CPO Recommendations first (numbered). Cover recruiting pipeline, time-to-hire, offer acceptance and QoH.",
}

def _pick_primary_cxo(blob: str) -> str:
    txt = (blob or "").lower()
    scores = {k: 0 for k in PRIMARY_KEYWORDS.keys()}
    for role, kws in PRIMARY_KEYWORDS.items():
        for kw in kws:
            if kw in txt:
                scores[role] += 1
    role = max(scores.items(), key=lambda kv: kv[1])[0] if scores else "CFO"
    return role or "CFO"

def _compose_premium_system_prompt(primary_role: str) -> str:
    base = (
        "You are CAIO Premium assistant. Be concise, actionable, and executive-friendly. "
        "Use numbered steps for plans and bullets for insights. If a document excerpt is provided, ground your answer in it. "
        "After the primary CXO block, add short notes from other CXOs when relevant."
    )
    hint = MODEL_HINT_MAP.get(primary_role, "")
    return f"{base}\nPrimary persona: {primary_role}. {hint}".strip()
# ====== end PRIMARY CXO ROUTING ======

def _llm_chat(messages: List[dict]) -> str:
    """
    messages: [{role:'system'|'user'|'assistant', content:str}, ...]
    """
    providers = _pick_provider(is_paid=True)  # use Pro provider quality for premium
    import requests
    if "openai" in providers:
        key = os.getenv("OPENAI_API_KEY")
        headers = {"Authorization": f"Bearer {key}", "Content-Type":"application/json"}
        if os.getenv("OPENAI_ORG"): headers["OpenAI-Organization"] = os.getenv("OPENAI_ORG")
        if os.getenv("OPENAI_PROJECT"): headers["OpenAI-Project"] = os.getenv("OPENAI_PROJECT")
        data = {"model":"gpt-4o-mini","messages":messages,"temperature":0.3}
        r = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, data=json.dumps(data), timeout=60)
        if r.status_code < 400:
            return r.json()["choices"][0]["message"]["content"]
    if "openrouter" in providers:
        key = os.getenv("OPENROUTER_API_KEY")
        headers = {"Authorization": f"Bearer {key}","HTTP-Referer":"https://caio-frontend.vercel.app","X-Title":"CAIO","Content-Type":"application/json"}
        data = {"model":"openai/gpt-4o-mini","messages":messages,"temperature":0.3}
        r = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, data=json.dumps(data), timeout=60)
        if r.status_code < 400:
            return r.json()["choices"][0]["message"]["content"]
    return "I’m online but couldn’t reach the model provider just now. Please try again in a bit."

def _ensure_session(db: Session, user_id: int, session_id: Optional[int], title_hint: Optional[str]) -> ChatSession:
    if session_id:
        sess = db.query(ChatSession).filter(ChatSession.id==session_id, ChatSession.user_id==user_id).first()
        if sess: return sess
    title = (title_hint or "New conversation")[:255]
    sess = ChatSession(user_id=user_id, title=title, created_at=datetime.utcnow())
    db.add(sess); db.commit(); db.refresh(sess)
    return sess

def _save_msg(db: Session, session_id: int, role: str, content: str) -> ChatMessage:
    m = ChatMessage(session_id=session_id, role=role, content=content, created_at=datetime.utcnow())
    db.add(m); db.commit(); db.refresh(m)
    return m

def _history(db: Session, session_id: int, user_id: int, limit: int = 20) -> List[ChatMessage]:
    q = (
        db.query(ChatMessage)
        .join(ChatSession, ChatMessage.session_id==ChatSession.id)
        .filter(ChatSession.id==session_id, ChatSession.user_id==user_id)
        .order_by(ChatMessage.id.desc())
        .limit(limit)
    )
    return list(reversed(q.all()))

@app.post("/api/chat/send")
async def chat_send(
    request: Request,
    message: str = Form(...),
    session_id: Optional[int] = Form(None),
    file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _is_premium_or_admin(current_user):
        raise HTTPException(status_code=403, detail="Chat is a Premium feature. Contact us for access.")
    db.add(UsageLog(user_id=getattr(current_user, "id", 0) or 0, timestamp=datetime.utcnow(), endpoint="chat", status="ok"))
    db.commit()
    sess = _ensure_session(db, getattr(current_user, "id", 0) or 0, session_id, title_hint=message[:60])
    context = ""
    try:
        context = (await _extract_text(file)).strip() if file else ""
    except Exception as e:
        logger.warning("File read failed: %s", e); context = ""
    primary_role = _pick_primary_cxo(f"{context}\n\n{message}")
    sys_prompt = _compose_premium_system_prompt(primary_role)
    msgs: List[dict] = [{"role":"system","content":sys_prompt}]
    for m in _history(db, sess.id, getattr(current_user, "id", 0) or 0, limit=20):
        msgs.append({"role": m.role, "content": m.content})
    if context:
        msgs.append({"role":"system","content": f"[DOCUMENT EXCERPT]\n{context[:8000]}"})
    msgs.append({"role":"user","content": message})
    _save_msg(db, sess.id, "user", message)
    reply = _llm_chat(msgs)[:120000]
    _save_msg(db, sess.id, "assistant", reply)
    return {"session_id": sess.id, "assistant": reply}

@app.get("/api/chat/history")
def chat_history(
    session_id: int = Query(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _is_premium_or_admin(current_user):
        raise HTTPException(status_code=403, detail="Chat is a Premium feature.")
    sess = db.query(ChatSession).filter(
        ChatSession.id==session_id, ChatSession.user_id==getattr(current_user, "id", 0)
    ).first()
    if not sess:
        raise HTTPException(status_code=404, detail="Session not found")
    items = [{
        "id": m.id, "role": m.role, "content": m.content, "created_at": m.created_at.isoformat()+"Z"
    } for m in _history(db, session_id, getattr(current_user,"id",0), limit=100)]
    return {"session_id": session_id, "messages": items, "title": sess.title}

@app.get("/api/chat/sessions")
def chat_sessions(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not _is_premium_or_admin(current_user):
        raise HTTPException(status_code=403, detail="Chat is a Premium feature.")
    rows = (
        db.query(ChatSession)
        .filter(ChatSession.user_id==getattr(current_user,"id",0))
        .order_by(ChatSession.id.desc())
        .limit(50)
        .all()
    )
    return [{"id": s.id, "title": s.title or f"Chat {s.id}", "created_at": s.created_at.isoformat()+"Z"} for s in rows]

# ---------- Admin Usage (analytics) ----------
def _require_admin(user: User):
    if _user_tier(user) != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

@app.get("/api/admin/usage")
def admin_usage(
    days: int = Query(14, ge=1, le=90),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin(current_user)

    now = datetime.utcnow()
    since = now - timedelta(days=days - 1)

    if db.bind.dialect.name == "postgresql":
        daycol = func.date_trunc("day", UsageLog.timestamp)
    else:  # sqlite
        daycol = func.strftime("%Y-%m-%d", UsageLog.timestamp)

    rows = (
        db.query(daycol.label("day"), UsageLog.endpoint, func.count(UsageLog.id).label("count"))
        .filter(UsageLog.timestamp >= since)
        .group_by("day", "endpoint")
        .order_by("day")
        .all()
    )

    def _day_key(v):
        if isinstance(v, str): return v[:10]
        try: return v.date().isoformat()
        except Exception: return str(v)[:10]

    day_keys = [(since + timedelta(days=i)).date().isoformat() for i in range(days)]
    series_map = {d: {"day": d, "analyze": 0, "chat": 0, "total": 0} for d in day_keys}
    totals = {"analyze": 0, "chat": 0, "total": 0}

    for day, endpoint, count in rows:
        dkey = _day_key(day)
        if dkey not in series_map:
            series_map[dkey] = {"day": dkey, "analyze": 0, "chat": 0, "total": 0}
        if endpoint == "chat":
            series_map[dkey]["chat"] += int(count); totals["chat"] += int(count)
        else:
            series_map[dkey]["analyze"] += int(count); totals["analyze"] += int(count)
        series_map[dkey]["total"] = series_map[dkey]["analyze"] + series_map[dkey]["chat"]

    totals["total"] = totals["analyze"] + totals["chat"]
    series = [series_map[d] for d in day_keys]

    return {
        "from": since.isoformat() + "Z",
        "to": now.isoformat() + "Z",
        "days": days,
        "series": series,
        "totals": totals,
    }

@app.get("/api/admin/users")
def admin_users(
    days: int = Query(14, ge=1, le=90),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin(current_user)

    since = datetime.utcnow() - timedelta(days=days - 1)

    q = (
        db.query(User.email.label("email"), func.count(UsageLog.id).label("count"))
        .join(UsageLog, UsageLog.user_id == User.id)
        .filter(UsageLog.timestamp >= since)
        .group_by(User.email)
        .order_by(func.count(UsageLog.id).desc())
        .limit(limit)
    )
    rows = q.all()

    flags_map = {}
    if rows:
        emails = [r.email for r in rows]
        user_rows = db.query(User).filter(User.email.in_(emails)).all()
        for u in user_rows:
            flags_map[u.email] = {
                "is_admin": bool(getattr(u, "is_admin", False)),
                "is_paid": bool(getattr(u, "is_paid", False)),
            }

    out = []
    for r in rows:
        flags = flags_map.get(r.email, {"is_admin": False, "is_paid": False})
        out.append({
            "email": r.email,
            "count": int(r.count),
            "is_admin": flags["is_admin"],
            "is_paid": flags["is_paid"],
        })

    return {
        "from": (datetime.utcnow() - timedelta(days=days - 1)).isoformat() + "Z",
        "to": datetime.utcnow().isoformat() + "Z",
        "days": days,
        "top": out,
    }

# ---------- Admin Users Roster (list & mutate) ----------
from pydantic import BaseModel  # add at top if not present

class SetPaidPayload(BaseModel):
    email: str
    is_paid: bool

@app.get("/api/admin/users/roster")
@app.get("/api/admin/users/roster/")  # accept trailing slash too
def admin_users_roster(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    q: Optional[str] = Query(None, description="search by email"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin(current_user)

    base = db.query(User)
    if q:
        like = f"%{q.lower()}%"
        base = base.filter(func.lower(User.email).like(like))

    total = base.count()
    rows = (
        base.order_by(User.created_at.desc() if hasattr(User, "created_at") else User.email.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
    )

    user_ids = [getattr(u, "id", 0) for u in rows if getattr(u, "id", 0)]
    stats: dict[int, dict] = {}
    if user_ids:
        now = datetime.utcnow()
        since1 = now - timedelta(days=1)
        since7 = now - timedelta(days=7)

        last_rows = (
            db.query(UsageLog.user_id, func.max(UsageLog.timestamp))
              .filter(UsageLog.user_id.in_(user_ids))
              .group_by(UsageLog.user_id)
              .all()
        )
        for uid, last_ts in last_rows:
            stats[int(uid)] = {"last_seen": (last_ts.isoformat() + "Z") if last_ts else None}

        r24 = (
            db.query(UsageLog.user_id, func.count(UsageLog.id))
              .filter(UsageLog.user_id.in_(user_ids), UsageLog.timestamp >= since1)
              .group_by(UsageLog.user_id)
              .all()
        )
        for uid, c in r24:
            stats.setdefault(int(uid), {})["count_24h"] = int(c)

        r24a = (
            db.query(UsageLog.user_id, func.count(UsageLog.id))
              .filter(
                  UsageLog.user_id.in_(user_ids),
                  UsageLog.timestamp >= since1,
                  UsageLog.endpoint == "analyze",
              )
              .group_by(UsageLog.user_id)
              .all()
        )
        for uid, c in r24a:
            stats.setdefault(int(uid), {})["analyze_24h"] = int(c)

        r24c = (
            db.query(UsageLog.user_id, func.count(UsageLog.id))
              .filter(
                  UsageLog.user_id.in_(user_ids),
                  UsageLog.timestamp >= since1,
                  UsageLog.endpoint == "chat",
              )
              .group_by(UsageLog.user_id)
              .all()
        )
        for uid, c in r24c:
            stats.setdefault(int(uid), {})["chat_24h"] = int(c)

        r7 = (
            db.query(UsageLog.user_id, func.count(UsageLog.id))
              .filter(UsageLog.user_id.in_(user_ids), UsageLog.timestamp >= since7)
              .group_by(UsageLog.user_id)
              .all()
        )
        for uid, c in r7:
            stats.setdefault(int(uid), {})["count_7d"] = int(c)

    items = []
    for u in rows:
        email = (u.email or "").lower()
        tier = "admin" if email in ADMIN_EMAILS else (
            "premium" if email in PREMIUM_EMAILS else (
                "pro" if bool(getattr(u, "is_paid", False)) else "demo"
            )
        )
        s = stats.get(getattr(u, "id", 0), {})
        created = getattr(u, "created_at", None)
        items.append({
            "email": u.email,
            "tier": tier,
            "is_admin": bool(email in ADMIN_EMAILS or getattr(u, "is_admin", False)),
            "is_paid": bool(getattr(u, "is_paid", False) or email in ADMIN_EMAILS or email in PREMIUM_EMAILS),
            "managed": "env" if (email in ADMIN_EMAILS or email in PREMIUM_EMAILS) else "db",
            "created_at": created.isoformat() + "Z" if created else None,
            "last_seen": s.get("last_seen"),
            "count_24h": s.get("count_24h", 0),
            "analyze_24h": s.get("analyze_24h", 0),
            "chat_24h": s.get("chat_24h", 0),
            "count_7d": s.get("count_7d", 0),
            "can_toggle_paid": email not in ADMIN_EMAILS and email not in PREMIUM_EMAILS,
        })

    return {"page": page, "page_size": page_size, "total": total, "items": items}

@app.post("/api/admin/users/set-paid")
def admin_users_set_paid(
    payload: SetPaidPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin(current_user)

    email_l = (payload.email or "").strip().lower()
    if not email_l:
        raise HTTPException(status_code=400, detail="email required")

    if email_l in ADMIN_EMAILS or email_l in PREMIUM_EMAILS:
        raise HTTPException(status_code=400, detail="This user is managed via env (ADMIN_EMAILS/PREMIUM_EMAILS)." )

    u = db.query(User).filter(func.lower(User.email) == email_l).first()
    if not u:
        raise HTTPException(status_code=404, detail="User not found")

    if not hasattr(u, "is_paid"):
        raise HTTPException(status_code=500, detail="User model missing 'is_paid' column")

    u.is_paid = bool(payload.is_paid)
    db.add(u); db.commit(); db.refresh(u)
    return {"email": u.email, "is_paid": bool(u.is_paid)}

# ---------- Admin: user summary & roster (alt view) ----------
from typing import Optional, Dict, List  # re-import safe
from pydantic import BaseModel  # ensure imported once
from sqlalchemy import func

USD_PER_TOKEN_FALLBACK = 0.000002  # used only if you don't log cost_usd

def _tier_for(email: str, is_paid: bool) -> str:
    e = (email or "").lower()
    if e in ADMIN_EMAILS or e in PREMIUM_EMAILS:
        return "premium"
    if is_paid:
        return "pro"
    return "demo"

@app.get("/api/admin/users/summary")
def admin_users_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin(current_user)

    users: List[User] = db.query(User).all()
    total = len(users)
    demo = pro = premium = 0
    for u in users:
        tier = _tier_for(getattr(u, "email", "") or "", bool(getattr(u, "is_paid", False)))
        if tier == "premium":
            premium += 1
        elif tier == "pro":
            pro += 1
        else:
            demo += 1

    return {"total_users": total, "demo": demo, "pro": pro, "premium": premium}

@app.get("/api/admin/users/roster")
def admin_users_roster_view(
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=200),
    q: Optional[str] = Query(None, description="search by email"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_admin(current_user)

    base = db.query(User)
    if q:
        like = f"%{q.lower()}%"
        base = base.filter(func.lower(User.email).like(like))

    total = base.count()
    users: List[User] = (
        base.order_by(User.created_at.desc() if hasattr(User, "created_at") else User.email.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
    )

    ids = [getattr(u, "id", 0) for u in users if getattr(u, "id", 0)]
    stats: Dict[int, Dict] = {int(uid): {} for uid in ids}

    if ids:
        last_rows = (
            db.query(UsageLog.user_id, func.max(UsageLog.timestamp))
              .filter(UsageLog.user_id.in_(ids))
              .group_by(UsageLog.user_id)
              .all()
        )
        for uid, last_ts in last_rows:
            stats[int(uid)]["last_seen"] = (last_ts.isoformat() + "Z") if last_ts else None

        try:
            sess_rows = (
                db.query(ChatSession.user_id, func.count(ChatSession.id))
                  .filter(ChatSession.user_id.in_(ids))
                  .group_by(ChatSession.user_id)
                  .all()
            )
            for uid, c in sess_rows:
                stats[int(uid)]["total_sessions"] = int(c or 0)
        except Exception:
            pass

        cost_col = getattr(UsageLog, "cost_usd", None)
        tokens_col = getattr(UsageLog, "total_tokens", None) or getattr(UsageLog, "tokens", None)

        if cost_col is not None:
            cost_rows = (
                db.query(UsageLog.user_id, func.coalesce(func.sum(cost_col), 0.0))
                  .filter(UsageLog.user_id.in_(ids))
                  .group_by(UsageLog.user_id)
                  .all()
            )
            for uid, cost in cost_rows:
                stats[int(uid)]["spend_usd"] = float(cost or 0.0)
        elif tokens_col is not None:
            tok_rows = (
                db.query(UsageLog.user_id, func.coalesce(func.sum(tokens_col), 0))
                  .filter(UsageLog.user_id.in_(ids))
                  .group_by(UsageLog.user_id)
                  .all()
            )
            for uid, toks in tok_rows:
                stats[int(uid)]["spend_usd"] = float(toks or 0) * USD_PER_TOKEN_FALLBACK

    items = []
    for u in users:
        uid = int(getattr(u, "id", 0) or 0)
        email = getattr(u, "email", "") or ""
        tier = _tier_for(email, bool(getattr(u, "is_paid", False)))
        created = getattr(u, "created_at", None)
        s = stats.get(uid, {}) if uid else {}
        items.append({
            "email": email,
            "tier": tier,
            "created_at": created.isoformat() + "Z" if created else None,
            "last_seen": s.get("last_seen"),
            "total_sessions": int(s.get("total_sessions", 0)),
            "spend_usd": float(s.get("spend_usd", 0.0)),
        })

    return {"page": page, "page_size": page_size, "total": total, "items": items}

# ---------- misc debug ----------
@app.get("/api/debug/ping-db")
def ping_db(db: Session = Depends(get_db)):
    try:
        count = db.query(User).count()
        return {"ok": True, "user_count": count}
    except Exception as e:
        if DEBUG:
            return JSONResponse({"ok": False, "detail": str(e)}, status_code=500)
        raise HTTPException(status_code=500, detail="DB error")

@app.get("/api/debug/routers")
def debug_routers():
    out = []
    for r in app.routes:
        methods = sorted(list(getattr(r, "methods", [])))
        path = getattr(r, "path", "")
        if path.startswith("/api/"):
            out.append({"path": path, "methods": methods})
    return out

# ---------- other routers ----------
try:
    from routes_public_config import router as public_cfg_router
    app.include_router(public_cfg_router)
except Exception as e:
    logger.warning(f"routes_public_config not loaded: {e}")

try:
    from payment_routes import router as payments_router
    app.include_router(payments_router)
except Exception as e:
    logger.warning(f"payment_routes not loaded: {e}")

# NEW: include the admin metrics router so /api/admin/metrics works
try:
    from admin_metrics_routes import router as admin_metrics_router
    app.include_router(admin_metrics_router)
except Exception as e:
    logger.warning(f"admin_metrics_routes not loaded: {e}")
